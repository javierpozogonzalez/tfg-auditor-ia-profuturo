import base64
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.platypus.flowables import KeepTogether

PROFUTURO_BLUE  = colors.HexColor("#003087")
PROFUTURO_CYAN  = colors.HexColor("#00AEEF")
ACCENT_RED      = colors.HexColor("#DC0000")
ACCENT_ORANGE   = colors.HexColor("#FF6400")
ACCENT_YELLOW   = colors.HexColor("#FFC800")
ACCENT_SOFT     = colors.HexColor("#5A96FF")
TEXT_DARK       = colors.HexColor("#1A1A2E")
TEXT_BODY       = colors.HexColor("#2D2D2D")
BG_LIGHT        = colors.HexColor("#F4F7FB")
RULE_COLOR      = colors.HexColor("#D0DFF0")

PAGE_W, PAGE_H  = A4
MARGIN          = 22 * mm


def _build_styles():
    base = getSampleStyleSheet()

    styles = {
        "doc_title": ParagraphStyle(
            "doc_title",
            fontName="Helvetica-Bold",
            fontSize=22,
            textColor=PROFUTURO_BLUE,
            spaceAfter=4,
            alignment=TA_CENTER,
            leading=28,
        ),
        "doc_subtitle": ParagraphStyle(
            "doc_subtitle",
            fontName="Helvetica",
            fontSize=10,
            textColor=PROFUTURO_CYAN,
            spaceAfter=2,
            alignment=TA_CENTER,
            leading=14,
        ),
        "h1": ParagraphStyle(
            "h1",
            fontName="Helvetica-Bold",
            fontSize=14,
            textColor=PROFUTURO_BLUE,
            spaceBefore=14,
            spaceAfter=4,
            leading=18,
            borderPad=0,
        ),
        "h2": ParagraphStyle(
            "h2",
            fontName="Helvetica-Bold",
            fontSize=12,
            textColor=PROFUTURO_CYAN,
            spaceBefore=10,
            spaceAfter=3,
            leading=16,
        ),
        "h3": ParagraphStyle(
            "h3",
            fontName="Helvetica-BoldOblique",
            fontSize=11,
            textColor=TEXT_DARK,
            spaceBefore=8,
            spaceAfter=2,
            leading=14,
        ),
        "body": ParagraphStyle(
            "body",
            fontName="Helvetica",
            fontSize=10,
            textColor=TEXT_BODY,
            spaceAfter=5,
            leading=15,
            alignment=TA_JUSTIFY,
        ),
        "bullet": ParagraphStyle(
            "bullet",
            fontName="Helvetica",
            fontSize=10,
            textColor=TEXT_BODY,
            spaceAfter=3,
            leading=14,
            leftIndent=22,
            bulletIndent=8,
            bulletFontName="Helvetica-Bold",
            bulletFontSize=10,
            bulletColor=PROFUTURO_CYAN,
        ),
        "bullet2": ParagraphStyle(
            "bullet2",
            fontName="Helvetica",
            fontSize=9.5,
            textColor=TEXT_BODY,
            spaceAfter=2,
            leading=13,
            leftIndent=40,
            bulletIndent=26,
            bulletFontName="Helvetica",
            bulletFontSize=9,
            bulletColor=PROFUTURO_CYAN,
        ),
        "kpi_label": ParagraphStyle(
            "kpi_label",
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=PROFUTURO_BLUE,
            alignment=TA_CENTER,
            leading=12,
        ),
        "kpi_value": ParagraphStyle(
            "kpi_value",
            fontName="Helvetica-Bold",
            fontSize=18,
            textColor=PROFUTURO_CYAN,
            alignment=TA_CENTER,
            leading=22,
        ),
        "footer": ParagraphStyle(
            "footer",
            fontName="Helvetica",
            fontSize=8,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ),
        "alert_title": ParagraphStyle(
            "alert_title",
            fontName="Helvetica-Bold",
            fontSize=20,
            textColor=ACCENT_RED,
            alignment=TA_CENTER,
            spaceAfter=6,
            leading=26,
        ),
        "alert_body": ParagraphStyle(
            "alert_body",
            fontName="Helvetica",
            fontSize=10,
            textColor=TEXT_BODY,
            spaceAfter=5,
            leading=15,
        ),
        "alert_section": ParagraphStyle(
            "alert_section",
            fontName="Helvetica-Bold",
            fontSize=11,
            textColor=TEXT_DARK,
            spaceBefore=10,
            spaceAfter=4,
            leading=15,
        ),
    }
    return styles


def _header_footer(canvas, doc):
    canvas.saveState()
    w, h = A4

    logo_path = str(Path(__file__).resolve().parent.parent.parent / "public" / "logo.png")
    if os.path.exists(logo_path):
        canvas.drawImage(logo_path, MARGIN, h - 14 * mm, width=30 * mm, height=12 * mm,
                         preserveAspectRatio=True, mask="auto")

    canvas.setStrokeColor(PROFUTURO_BLUE)
    canvas.setLineWidth(1.5)
    canvas.line(MARGIN, h - 16 * mm, w - MARGIN, h - 16 * mm)

    canvas.setStrokeColor(RULE_COLOR)
    canvas.setLineWidth(0.5)
    canvas.line(MARGIN, 13 * mm, w - MARGIN, 13 * mm)

    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(colors.HexColor("#888888"))
    canvas.drawString(MARGIN, 9 * mm,
                      f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Confidencial — Solo uso interno")
    canvas.drawRightString(w - MARGIN, 9 * mm, f"Pagina {doc.page}")

    canvas.restoreState()


def _parse_markdown_to_story(text: str, styles: dict) -> list:
    story = []
    lines = text.splitlines()
    i = 0

    kpi_buffer = []

    def flush_kpis():
        if not kpi_buffer:
            return
        col_data = []
        for label, value in kpi_buffer:
            cell = [
                Paragraph(value, styles["kpi_value"]),
                Paragraph(label, styles["kpi_label"]),
            ]
            col_data.append(cell)

        num_cols = min(len(col_data), 4)
        rows = [col_data[j:j+num_cols] for j in range(0, len(col_data), num_cols)]

        for row in rows:
            while len(row) < num_cols:
                row.append([""])
            col_w = (PAGE_W - 2 * MARGIN) / num_cols
            tbl = Table([row], colWidths=[col_w] * num_cols, rowHeights=[22 * mm])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), BG_LIGHT),
                ("ROUNDEDCORNERS", [6]),
                ("BOX", (0, 0), (-1, -1), 0.5, RULE_COLOR),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, RULE_COLOR),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 4 * mm))
        kpi_buffer.clear()

    def _render_md_table(table_lines: list[str]) -> None:
        rows_raw = []
        for tl in table_lines:
            if re.match(r"^\|[-|:\s]+\|$", tl):
                continue
            cells = [c.strip() for c in tl.strip("|").split("|")]
            rows_raw.append(cells)
        if not rows_raw:
            return
        max_cols = max(len(r) for r in rows_raw)
        for r in rows_raw:
            while len(r) < max_cols:
                r.append("")

        header_style = ParagraphStyle(
            "th", fontName="Helvetica-Bold", fontSize=9,
            textColor=colors.white, alignment=TA_CENTER, leading=12,
        )
        cell_style = ParagraphStyle(
            "td", fontName="Helvetica", fontSize=9,
            textColor=TEXT_BODY, alignment=TA_LEFT, leading=12,
        )
        table_data = []
        for ri, row in enumerate(rows_raw):
            st = header_style if ri == 0 else cell_style
            table_data.append([Paragraph(_inline_markdown(c), st) for c in row])

        col_w = (PAGE_W - 2 * MARGIN) / max_cols
        tbl = Table(table_data, colWidths=[col_w] * max_cols, repeatRows=1)
        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PROFUTURO_BLUE),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_LIGHT]),
            ("BOX", (0, 0), (-1, -1), 0.5, RULE_COLOR),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, RULE_COLOR),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ])
        tbl.setStyle(ts)
        story.append(tbl)
        story.append(Spacer(1, 3 * mm))

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if stripped.startswith("|"):
            flush_kpis()
            table_block = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_block.append(lines[i].strip())
                i += 1
            _render_md_table(table_block)
            continue

        elif re.match(r'^#{1,6} ', stripped):
            flush_kpis()
            hm = re.match(r'^(#{1,6})\s+(.+)', stripped)
            level = len(hm.group(1))
            text_h = hm.group(2).strip()
            if level == 1:
                story.append(Spacer(1, 3 * mm))
                story.append(Paragraph(text_h, styles["h1"]))
                story.append(HRFlowable(width="100%", thickness=1.5,
                                        color=PROFUTURO_BLUE, spaceAfter=3))
            elif level == 2:
                story.append(Spacer(1, 2 * mm))
                story.append(Paragraph(text_h, styles["h2"]))
                story.append(HRFlowable(width="100%", thickness=0.5,
                                        color=PROFUTURO_CYAN, spaceAfter=2))
            else:
                story.append(Paragraph(text_h, styles["h3"]))

        elif line.startswith("    - ") or line.startswith("    * ") or line.startswith("  - ") or line.startswith("  * "):
            flush_kpis()
            content = _inline_markdown(stripped[2:])
            story.append(Paragraph(f"<bullet>\u2013</bullet> {content}", styles["bullet2"]))

        elif stripped.startswith("- ") or stripped.startswith("* "):
            flush_kpis()
            content = stripped[2:]
            content = _inline_markdown(content)
            story.append(Paragraph(f"<bullet>\u2022</bullet> {content}", styles["bullet"]))

        elif re.match(r"^\d+\.\s", stripped):
            flush_kpis()
            content = re.sub(r"^\d+\.\s", "", stripped)
            content = _inline_markdown(content)
            num = re.match(r"^(\d+)\.", stripped).group(1)
            story.append(Paragraph(f"<bullet>{num}.</bullet> {content}", styles["bullet"]))

        elif stripped.startswith("---") or stripped.startswith("___"):
            flush_kpis()
            story.append(HRFlowable(width="100%", thickness=0.5,
                                    color=RULE_COLOR, spaceBefore=4, spaceAfter=4))

        elif re.match(r"^[A-Za-záéíóúÁÉÍÓÚñÑ ]+:\s+[\d\+\-\.,]+", stripped):
            parts = stripped.split(":", 1)
            if len(parts) == 2:
                kpi_buffer.append((parts[0].strip(), parts[1].strip()))

        elif stripped == "":
            flush_kpis()
            story.append(Spacer(1, 3 * mm))

        else:
            flush_kpis()
            content = _inline_markdown(stripped)
            story.append(Paragraph(content, styles["body"]))

        i += 1

    flush_kpis()
    return story


def _inline_markdown(text: str) -> str:
    text = re.sub(r"\*\*\*(.+?)\*\*\*", r"<b><i>\1</i></b>", text)
    text = re.sub(r"\*\*(.+?)\*\*",     r"<b>\1</b>",         text)
    text = re.sub(r"\*(.+?)\*",         r"<i>\1</i>",         text)
    text = re.sub(r"`(.+?)`",           r"<font name='Courier'>\1</font>", text)
    text = text.replace("&", "&amp;")
    return text


def generate_report_pdf(text: str, title: str) -> str:
    title_clean = title.replace("_", " ")
    styles = _build_styles()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp_path = tmp.name

    doc = SimpleDocTemplate(
        tmp_path,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=22 * mm,
        bottomMargin=20 * mm,
        title=title_clean,
        author="ProFuturo Auditor IA",
    )

    story = []

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(title_clean, styles["doc_title"]))
    story.append(Paragraph("Informe generado automaticamente por el Auditor IA de ProFuturo", styles["doc_subtitle"]))
    story.append(Spacer(1, 2 * mm))
    story.append(HRFlowable(width="60%", thickness=2, color=PROFUTURO_CYAN,
                             hAlign="CENTER", spaceAfter=6))
    story.append(Spacer(1, 4 * mm))

    story += _parse_markdown_to_story(text, styles)

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    with open(tmp_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("utf-8")
    os.unlink(tmp_path)
    return encoded


def generate_report_excel(data: dict, title: str) -> str:
    """Genera un .xlsx profesional con portada, 4 hojas de datos y gráficos. Devuelve base64."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Paleta corporativa ────────────────────────────────────────────
    BLUE_DARK  = "003087"
    BLUE_MED   = "0066CC"
    CYAN_COL   = "00AEEF"
    WHITE      = "FFFFFF"
    ALT_ROW    = "F8F9FA"
    BORDER_CLR = "E0E0E0"
    GREEN_COL  = "007A33"
    RED_COL    = "CC0000"
    TEXT_DARK  = "1A1A2E"
    ORANGE     = "E67E00"

    HDR_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
    HDR_FILL = PatternFill("solid", fgColor=BLUE_MED)
    ALT_FILL = PatternFill("solid", fgColor=ALT_ROW)
    CTR      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT     = Alignment(horizontal="left",   vertical="center")
    RIGHT    = Alignment(horizontal="right",  vertical="center")

    def _thin_border():
        s = Side(style="thin", color=BORDER_CLR)
        return Border(left=s, right=s, top=s, bottom=s)

    def _style_header_row(ws, headers, row=1):
        ws.row_dimensions[row].height = 22
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = CTR
            c.border = _thin_border()
        ws.freeze_panes = ws.cell(row + 1, 1)
        ws.sheet_view.showGridLines = False

    def _data_cell(ws, row, col, val, alt=False, align=CTR, color=TEXT_DARK, bold=False):
        c = ws.cell(row, col, val)
        c.font   = Font(name="Arial", size=10, color=color, bold=bold)
        c.alignment = align
        c.border = _thin_border()
        if alt:
            c.fill = ALT_FILL

    def _add_bar_chart(ws, data_range, categories_range, title, position, horizontal=True):
        try:
            chart = BarChart()
            chart.type     = "bar" if horizontal else "col"
            chart.style    = 10
            chart.title    = title
            chart.y_axis.title = "Usuarios" if horizontal else "Cantidad"
            chart.x_axis.title = "Cantidad" if horizontal else "Usuarios"
            chart.height   = 10
            chart.width    = 15
            d_ref = Reference(ws, min_col=data_range[0],       min_row=data_range[1],       max_row=data_range[2])
            c_ref = Reference(ws, min_col=categories_range[0], min_row=categories_range[1], max_row=categories_range[2])
            chart.add_data(d_ref, titles_from_data=True)
            chart.set_categories(c_ref)
            ws.add_chart(chart, position)
        except Exception:
            pass

    wb = Workbook()
    wb.remove(wb.active)
    community_name = title.replace("_", " ").strip()

    # ── PORTADA ───────────────────────────────────────────────────────
    ws_cv = wb.create_sheet("Portada")
    ws_cv.sheet_view.showGridLines = False
    ws_cv.column_dimensions["A"].width = 3
    ws_cv.column_dimensions["B"].width = 58
    ws_cv.column_dimensions["C"].width = 18

    # Banda azul superior (filas 1–3)
    blue_fill = PatternFill("solid", fgColor=BLUE_DARK)
    for r in range(1, 4):
        ws_cv.row_dimensions[r].height = 14
        for c in range(1, 9):
            ws_cv.cell(r, c).fill = blue_fill

    # Logo
    logo_path = str(Path(__file__).resolve().parent.parent.parent / "public" / "logo.png")
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            xl_logo = XLImage(logo_path)
            xl_logo.width  = 130
            xl_logo.height = 48
            ws_cv.add_image(xl_logo, "B2")
        except Exception:
            pass

    ws_cv.row_dimensions[4].height = 22

    ws_cv.row_dimensions[5].height = 42
    c = ws_cv["B5"]
    c.value = f"Reporte de Actividad — {community_name}"
    c.font  = Font(name="Arial", bold=True, size=20, color=BLUE_DARK)
    c.alignment = LEFT

    ws_cv.row_dimensions[6].height = 24
    c = ws_cv["B6"]
    c.value = "Análisis de Participación y Engagement"
    c.font  = Font(name="Arial", size=13, color=BLUE_MED)
    c.alignment = LEFT

    ws_cv.row_dimensions[7].height = 10
    # Línea separadora (fila 8 con fondo azul medio, altura 2)
    ws_cv.row_dimensions[8].height = 2
    for col in range(1, 9):
        ws_cv.cell(8, col).fill = PatternFill("solid", fgColor=BLUE_MED)

    ws_cv.row_dimensions[9].height = 14
    c = ws_cv["B9"]
    c.value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
    c.font  = Font(name="Arial", size=10, color="555555")
    c.alignment = LEFT

    ws_cv.row_dimensions[10].height = 14
    c = ws_cv["B10"]
    c.value = "Generado por: Auditor IA — ProFuturo"
    c.font  = Font(name="Arial", italic=True, size=10, color=CYAN_COL)
    c.alignment = LEFT

    ws_cv.row_dimensions[12].height = 16
    c = ws_cv["B12"]
    c.value = "Contenido del reporte:"
    c.font  = Font(name="Arial", bold=True, size=11, color=BLUE_DARK)

    index_desc = [
        ("KPIs",             "Indicadores clave de participación y engagement"),
        ("Ranking Usuarios", "Top usuarios por volumen de publicaciones"),
        ("Conectores",       "Top usuarios por diversidad de participación"),
        ("Tendencias",       "Temas más debatidos y su actividad reciente"),
    ]
    for idx, (sname, sdesc) in enumerate(index_desc, 13):
        ws_cv.row_dimensions[idx].height = 16
        c = ws_cv[f"B{idx}"]
        c.value = f"  •  {sname}: {sdesc}"
        c.font  = Font(name="Arial", size=10, color="333333")

    # ── KPIs ──────────────────────────────────────────────────────────
    kpi_data = data.get("KPIs", {})
    if kpi_data:
        ws = wb.create_sheet("KPIs")
        _style_header_row(ws, ["Métrica", "Valor Actual", "Mes Anterior", "Variación"])
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 22

        for ri, row_data in enumerate(kpi_data.get("rows", []), 2):
            alt   = ri % 2 == 0
            name  = str(row_data[0]) if len(row_data) > 0 else ""
            val   = row_data[1]      if len(row_data) > 1 else ""
            var   = str(row_data[2]) if len(row_data) > 2 else "—"

            # Detectar dirección de variación
            var_color = TEXT_DARK
            if "+" in var and "(" in var:
                var_color = GREEN_COL
                var = "↗ " + var
            elif "-" in var and "(" in var:
                var_color = RED_COL
                var = "↘ " + var

            _data_cell(ws, ri, 1, name,  alt=alt, align=LEFT)
            _data_cell(ws, ri, 2, val,   alt=alt, align=RIGHT)
            _data_cell(ws, ri, 3, "—",   alt=alt, align=CTR)
            _data_cell(ws, ri, 4, var,   alt=alt, align=CTR, color=var_color, bold=(var_color != TEXT_DARK))

    # ── RANKING USUARIOS ─────────────────────────────────────────────
    rank_data = data.get("Ranking_Usuarios", {})
    rank_rows = rank_data.get("rows", [])
    medals    = ["🥇", "🥈", "🥉"]

    if rank_rows:
        ws = wb.create_sheet("Ranking Usuarios")
        _style_header_row(ws, ["#", "Usuario", "Publicaciones", "Puntuación", "Rol"])
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16

        max_posts = int(rank_rows[0][2]) if rank_rows and len(rank_rows[0]) > 2 else 1

        for ri, row_data in enumerate(rank_rows, 2):
            alt  = ri % 2 == 0
            pos  = int(row_data[0]) if len(row_data) > 0 else ri - 1
            name = str(row_data[1]) if len(row_data) > 1 else ""
            posts = int(row_data[2]) if len(row_data) > 2 else 0
            discs = int(row_data[3]) if len(row_data) > 3 else 0
            score = round(min(10.0, posts / max(max_posts, 1) * 10), 1)
            score_str = f"{score:.1f} / 10"
            medal = medals[pos - 1] if pos <= 3 else str(pos)
            rol   = "Líder" if pos == 1 else ("Experto" if pos <= 3 else ("Conector" if discs > 2 else "Participante"))

            _data_cell(ws, ri, 1, medal,     alt=alt, align=CTR)
            _data_cell(ws, ri, 2, name,      alt=alt, align=LEFT)
            _data_cell(ws, ri, 3, posts,     alt=alt, align=RIGHT)
            _data_cell(ws, ri, 4, score_str, alt=alt, align=CTR)
            _data_cell(ws, ri, 5, rol,       alt=alt, align=CTR)

        _add_bar_chart(ws,
            data_range=(3, 1, len(rank_rows) + 1),
            categories_range=(2, 2, len(rank_rows) + 1),
            title="Top Participantes",
            position="G2",
            horizontal=True,
        )

    # ── CONECTORES ───────────────────────────────────────────────────
    if rank_rows:
        ws = wb.create_sheet("Conectores")
        _style_header_row(ws, ["#", "Usuario", "Discusiones", "Publicaciones", "Rol"])
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 16

        sorted_by_disc = sorted(rank_rows, key=lambda r: int(r[3]) if len(r) > 3 else 0, reverse=True)
        for ri, row_data in enumerate(sorted_by_disc, 2):
            alt   = ri % 2 == 0
            pos   = ri - 1
            name  = str(row_data[1]) if len(row_data) > 1 else ""
            posts = int(row_data[2]) if len(row_data) > 2 else 0
            discs = int(row_data[3]) if len(row_data) > 3 else 0
            medal = medals[pos - 1] if pos <= 3 else str(pos)
            rol   = "Conector" if discs >= 5 else ("Experto" if discs >= 3 else "Participante")

            _data_cell(ws, ri, 1, medal, alt=alt, align=CTR)
            _data_cell(ws, ri, 2, name,  alt=alt, align=LEFT)
            _data_cell(ws, ri, 3, discs, alt=alt, align=RIGHT)
            _data_cell(ws, ri, 4, posts, alt=alt, align=RIGHT)
            _data_cell(ws, ri, 5, rol,   alt=alt, align=CTR)

        _add_bar_chart(ws,
            data_range=(3, 1, len(sorted_by_disc) + 1),
            categories_range=(2, 2, len(sorted_by_disc) + 1),
            title="Usuarios por Diversidad de Discusiones",
            position="G2",
            horizontal=True,
        )

    # ── TENDENCIAS ───────────────────────────────────────────────────
    tend_data = data.get("Tendencias", {})
    tend_rows = tend_data.get("rows", [])
    if tend_rows:
        ws = wb.create_sheet("Tendencias")
        _style_header_row(ws, ["Tema", "Posts Totales", "Posts 30 días", "% Reciente", "Tendencia"])
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 18

        for ri, row_data in enumerate(tend_rows, 2):
            alt    = ri % 2 == 0
            tema   = str(row_data[0]) if len(row_data) > 0 else ""
            total  = row_data[1]      if len(row_data) > 1 else 0
            recent = row_data[2]      if len(row_data) > 2 else 0
            pct    = str(row_data[3]) if len(row_data) > 3 else "0%"
            try:
                pct_val  = float(pct.replace("%", ""))
                trend    = "↗ Creciendo" if pct_val >= 50 else ("→ Estable" if pct_val >= 20 else "↘ Bajando")
                t_color  = GREEN_COL if pct_val >= 50 else (ORANGE if pct_val >= 20 else RED_COL)
            except Exception:
                trend, t_color = "—", TEXT_DARK

            _data_cell(ws, ri, 1, tema,   alt=alt, align=LEFT)
            _data_cell(ws, ri, 2, total,  alt=alt, align=RIGHT)
            _data_cell(ws, ri, 3, recent, alt=alt, align=RIGHT)
            _data_cell(ws, ri, 4, pct,    alt=alt, align=CTR)
            _data_cell(ws, ri, 5, trend,  alt=alt, align=CTR, color=t_color, bold=True)

        _add_bar_chart(ws,
            data_range=(2, 1, len(tend_rows) + 1),
            categories_range=(1, 2, len(tend_rows) + 1),
            title="Temas más activos",
            position="G2",
            horizontal=False,
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    with open(tmp_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("utf-8")
    os.unlink(tmp_path)
    return encoded


def generate_critical_alert_pdf(community: str, issue_summary: str, severity: str) -> str:
    styles = _build_styles()

    severity_upper = (severity or "MEDIA").upper()
    severity_lower = severity_upper.lower()

    severity_colors = {
        "critica": ACCENT_RED,
        "alta":    ACCENT_ORANGE,
        "media":   ACCENT_YELLOW,
        "baja":    ACCENT_SOFT,
    }
    sev_color = severity_colors.get(severity_lower, colors.grey)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp_path = tmp.name

    doc = SimpleDocTemplate(
        tmp_path,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=22 * mm,
        bottomMargin=20 * mm,
        title=f"Alerta {severity_upper} — {community}",
        author="ProFuturo Auditor IA",
    )

    story = []

    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(f"ALERTA {severity_upper}", styles["alert_title"]))

    badge_data = [[Paragraph(severity_upper, ParagraphStyle(
        "badge", fontName="Helvetica-Bold", fontSize=13,
        textColor=colors.white, alignment=TA_CENTER, leading=18,
    ))]]
    badge = Table(badge_data, colWidths=[50 * mm], rowHeights=[10 * mm])
    badge.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), sev_color),
        ("ROUNDEDCORNERS", [6]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    story.append(Table([[badge]], colWidths=[PAGE_W - 2 * MARGIN]))
    story.append(Spacer(1, 4 * mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=sev_color, spaceAfter=4))

    meta_data = [
        ["Comunidad",  community],
        ["Severidad",  severity_upper],
        ["Fecha",      datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ["Sistema",    "ProFuturo Auditor IA — Monitor Proactivo"],
    ]
    meta_style = ParagraphStyle("meta", fontName="Helvetica", fontSize=10,
                                textColor=TEXT_BODY, leading=14)
    meta_label = ParagraphStyle("meta_label", fontName="Helvetica-Bold", fontSize=10,
                                textColor=PROFUTURO_BLUE, leading=14)
    meta_rows = [[Paragraph(k, meta_label), Paragraph(v, meta_style)] for k, v in meta_data]
    meta_tbl = Table(meta_rows, colWidths=[45 * mm, PAGE_W - 2 * MARGIN - 45 * mm])
    meta_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BG_LIGHT),
        ("BOX", (0, 0), (-1, -1), 0.5, RULE_COLOR),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, RULE_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Detalle de incidencias detectadas", styles["alert_section"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=RULE_COLOR, spaceAfter=3))
    for line in issue_summary.strip().splitlines():
        if line.strip():
            story.append(Paragraph(f"<bullet>\u2022</bullet> {line.strip()}", styles["bullet"]))
    story.append(Spacer(1, 5 * mm))

    story.append(Paragraph("Acciones recomendadas", styles["alert_section"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=RULE_COLOR, spaceAfter=3))
    actions = [
        "Revisar inmediatamente el contenido y los usuarios involucrados.",
        "Aplicar moderación o restricciones de acceso si es necesario.",
        "Contactar a los administradores de la comunidad afectada.",
        "Documentar el incidente en el sistema de tickets interno.",
        "Realizar seguimiento de resolución en un plazo máximo de 24 horas.",
    ]
    for idx, action in enumerate(actions, 1):
        story.append(Paragraph(f"<bullet>{idx}.</bullet> {action}", styles["bullet"]))

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)

    with open(tmp_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("utf-8")
    os.unlink(tmp_path)
    return encoded
